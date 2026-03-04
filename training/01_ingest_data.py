import os
import sys
import openpyxl
import pandas as pd
import numpy as np
from supabase import create_client
from dotenv import load_dotenv

# Load .env from project root
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), '..', '.env'))

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY")

supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

PROCESS_FILE = os.path.join(os.path.dirname(__file__), '_h_batch_process_data.xlsx')
PRODUCTION_FILE = os.path.join(os.path.dirname(__file__), '_h_batch_production_data.xlsx')

PHASE_ORDER = {
    'Preparation': 1,
    'Granulation': 2,
    'Drying': 3,
    'Milling': 4,
    'Blending': 5,
    'Compression': 6,
    'Coating': 7,
    'Quality_Testing': 8
}

def load_production_data():
    print("Loading production data...")
    wb = openpyxl.load_workbook(PRODUCTION_FILE)
    ws = wb['BatchData']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, 1):
            row[h] = ws.cell(r, c).value
        rows.append(row)
    print(f"  Loaded {len(rows)} production records")
    return rows

def load_summary_data():
    print("Loading process summary data...")
    wb = openpyxl.load_workbook(PROCESS_FILE)
    ws = wb['Summary']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = {}
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, 1):
            row[h] = ws.cell(r, c).value
        rows[row['Batch_ID']] = row
    print(f"  Loaded {len(rows)} summary records")
    return rows

def load_phase_data(batch_id):
    wb = openpyxl.load_workbook(PROCESS_FILE)
    sheet_name = f'Batch_{batch_id}'
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, 1):
            row[h] = ws.cell(r, c).value
        rows.append(row)
    return rows

def compute_phase_aggregates(batch_id):
    rows = load_phase_data(batch_id)
    if not rows:
        return []

    # Group by phase
    phases = {}
    for row in rows:
        phase = row.get('Phase', 'Unknown')
        if phase not in phases:
            phases[phase] = []
        phases[phase].append(row)

    phase_records = []
    for phase_name, phase_rows in phases.items():
        power_vals = [r['Power_Consumption_kW'] for r in phase_rows
                      if r['Power_Consumption_kW'] is not None]
        temp_vals = [r['Temperature_C'] for r in phase_rows
                     if r['Temperature_C'] is not None]
        pressure_vals = [r['Pressure_Bar'] for r in phase_rows
                         if r['Pressure_Bar'] is not None]
        vibration_vals = [r['Vibration_mm_s'] for r in phase_rows
                          if r['Vibration_mm_s'] is not None and r['Vibration_mm_s'] >= 0]
        motor_vals = [r['Motor_Speed_RPM'] for r in phase_rows
                      if r['Motor_Speed_RPM'] is not None and r['Motor_Speed_RPM'] > 0]
        comp_vals = [r['Compression_Force_kN'] for r in phase_rows
                     if r['Compression_Force_kN'] is not None and r['Compression_Force_kN'] > 0]

        duration = len(phase_rows)
        avg_power = float(np.mean(power_vals)) if power_vals else 0
        energy_kwh = avg_power * duration / 60

        phase_records.append({
            'batch_id': batch_id,
            'phase_name': phase_name,
            'phase_order': PHASE_ORDER.get(phase_name, 99),
            'duration_minutes': duration,
            'avg_temperature': float(np.mean(temp_vals)) if temp_vals else None,
            'max_temperature': float(np.max(temp_vals)) if temp_vals else None,
            'avg_pressure': float(np.mean(pressure_vals)) if pressure_vals else None,
            'avg_power_kw': avg_power,
            'max_power_kw': float(np.max(power_vals)) if power_vals else None,
            'avg_vibration': float(np.mean(vibration_vals)) if vibration_vals else None,
            'max_motor_speed': float(np.max(motor_vals)) if motor_vals else None,
            'max_compression_force': float(np.max(comp_vals)) if comp_vals else None,
            'energy_kwh': float(energy_kwh),
            'anomaly_score': 0.0
        })

    return phase_records

def is_batch_feasible(row):
    try:
        return (
            row.get('Dissolution_Rate', 0) > 85 and
            row.get('Friability', 99) < 1.0 and
            row.get('Hardness', 0) > 50 and
            row.get('Disintegration_Time', 99) < 15 and
            195 <= row.get('Tablet_Weight', 0) <= 207 and
            85 <= row.get('Content_Uniformity', 0) <= 115
        )
    except:
        return False

def seed_batches():
    print("\nSeeding batches table...")
    production_data = load_production_data()
    summary_data = load_summary_data()

    batch_records = []
    for prod in production_data:
        batch_id = prod['Batch_ID']
        summary = summary_data.get(batch_id, {})

        # Compute total energy from summary
        avg_power = summary.get('Avg_Power_Consumption', 0) or 0
        duration = summary.get('Duration_Minutes', 0) or 0
        total_energy = avg_power * duration / 60

        record = {
            'batch_id': batch_id,
            'granulation_time': prod.get('Granulation_Time'),
            'binder_amount': prod.get('Binder_Amount'),
            'drying_temp': prod.get('Drying_Temp'),
            'drying_time': prod.get('Drying_Time'),
            'compression_force': prod.get('Compression_Force'),
            'machine_speed': prod.get('Machine_Speed'),
            'lubricant_conc': prod.get('Lubricant_Conc'),
            'moisture_content': prod.get('Moisture_Content'),
            'tablet_weight': prod.get('Tablet_Weight'),
            'hardness': prod.get('Hardness'),
            'friability': prod.get('Friability'),
            'disintegration_time': prod.get('Disintegration_Time'),
            'dissolution_rate': prod.get('Dissolution_Rate'),
            'content_uniformity': prod.get('Content_Uniformity'),
            'total_energy_kwh': round(float(total_energy), 2),
            'batch_duration_minutes': duration,
            'is_feasible': is_batch_feasible(prod)
        }
        batch_records.append(record)

    # Insert in batches of 20
    for i in range(0, len(batch_records), 20):
        chunk = batch_records[i:i+20]
        result = supabase.table('batches').upsert(chunk).execute()
        print(f"  Inserted batches {i+1}–{min(i+20, len(batch_records))}")

    feasible_count = sum(1 for r in batch_records if r['is_feasible'])
    print(f"  Total: {len(batch_records)} batches | Feasible: {feasible_count}")
    return batch_records

def seed_phase_sensors():
    print("\nSeeding phase_sensors table...")
    wb = openpyxl.load_workbook(PROCESS_FILE)
    batch_sheets = [s for s in wb.sheetnames if s.startswith('Batch_')]
    
    total_phases = 0
    for sheet_name in batch_sheets:
        batch_id = sheet_name.replace('Batch_', '')
        phase_records = compute_phase_aggregates(batch_id)
        if phase_records:
            supabase.table('phase_sensors').upsert(phase_records).execute()
            total_phases += len(phase_records)
        print(f"  {batch_id}: {len(phase_records)} phases seeded")

    print(f"  Total phase records: {total_phases}")

def seed_model_metadata():
    print("\nSeeding model metadata...")
    metadata = [
        {
            'model_name': 'gpr_dissolution',
            'trained_on_batches': 60,
            'avg_accuracy': 99.11,
            'mape': 0.89,
            'conformal_coverage': 95.3
        },
        {
            'model_name': 'gpr_friability',
            'trained_on_batches': 60,
            'avg_accuracy': 92.6,
            'mape': 7.4,
            'conformal_coverage': 95.0
        },
        {
            'model_name': 'gpr_hardness',
            'trained_on_batches': 60,
            'avg_accuracy': 95.7,
            'mape': 4.3,
            'conformal_coverage': 95.0
        },
        {
            'model_name': 'gpr_disintegration',
            'trained_on_batches': 60,
            'avg_accuracy': 93.4,
            'mape': 6.6,
            'conformal_coverage': 96.7
        },
        {
            'model_name': 'gpr_tablet_weight',
            'trained_on_batches': 60,
            'avg_accuracy': 99.58,
            'mape': 0.42,
            'conformal_coverage': 96.7
        },
        {
            'model_name': 'gpr_content_uniformity',
            'trained_on_batches': 60,
            'avg_accuracy': 99.04,
            'mape': 0.96,
            'conformal_coverage': 93.3
        }
    ]
    supabase.table('model_metadata').upsert(metadata).execute()
    print(f"  Seeded {len(metadata)} model records")

def verify_seeding():
    print("\nVerifying seeding...")
    batches = supabase.table('batches').select('batch_id, is_feasible').execute()
    phases = supabase.table('phase_sensors').select('id').execute()
    print(f"  Batches in DB: {len(batches.data)}")
    print(f"  Phase records in DB: {len(phases.data)}")
    feasible = [b for b in batches.data if b['is_feasible']]
    print(f"  Feasible batches: {len(feasible)}")
    print(f"  Feasible batch IDs: {[b['batch_id'] for b in feasible]}")

if __name__ == '__main__':
    print("=" * 50)
    print("BatchMind — Data Ingestion Script")
    print("=" * 50)

    try:
        seed_batches()
        seed_phase_sensors()
        seed_model_metadata()
        verify_seeding()
        print("\nDone. Database seeded successfully.")
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)