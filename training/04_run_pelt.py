import os
import sys
import json
import numpy as np
import joblib
import openpyxl
from supabase import create_client
from dotenv import load_dotenv
import ruptures as rpt

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), '..', '.env'))

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY")
supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'backend', 'ml')
PROCESS_FILE = os.path.join(os.path.dirname(__file__), '_h_batch_process_data.xlsx')

PHASE_ORDER = [
    'Preparation', 'Granulation', 'Drying', 'Milling',
    'Blending', 'Compression', 'Coating', 'Quality_Testing'
]

def load_batch_power(batch_id):
    wb = openpyxl.load_workbook(PROCESS_FILE)
    sheet_name = f'Batch_{batch_id}'
    if sheet_name not in wb.sheetnames:
        return None, None
    ws = wb[sheet_name]
    power, phases = [], []
    for r in range(2, ws.max_row + 1):
        p = ws.cell(r, 10).value  # Power_Consumption_kW col
        ph = ws.cell(r, 3).value  # Phase col
        if p is not None:
            power.append(max(0, float(p)))
            phases.append(ph)
    return np.array(power), phases

def get_all_batch_ids():
    result = supabase.table('batches').select('batch_id').execute()
    return [r['batch_id'] for r in result.data]

def compute_phase_fingerprints(all_phase_power):
    """Build reference fingerprint (mean ± std) per phase"""
    print("\nBuilding phase fingerprints...")
    fingerprints = {}

    for phase in PHASE_ORDER:
        curves = all_phase_power.get(phase, [])
        if not curves:
            continue

        # Normalize all curves to same length (50 points)
        normalized = []
        for curve in curves:
            if len(curve) < 2:
                continue
            indices = np.linspace(0, len(curve) - 1, 50)
            resampled = np.interp(indices, np.arange(len(curve)), curve)
            normalized.append(resampled)

        if not normalized:
            continue

        arr = np.array(normalized)
        mean_curve = arr.mean(axis=0)
        std_curve = arr.std(axis=0)

        fingerprints[phase] = {
            'mean': mean_curve.tolist(),
            'std': std_curve.tolist(),
            'upper': (mean_curve + 2 * std_curve).tolist(),
            'lower': np.maximum(0, mean_curve - 2 * std_curve).tolist(),
            'avg_energy_kwh': float(np.mean([
                np.mean(c) * len(c) / 60 for c in curves
            ])),
            'sample_count': len(normalized)
        }
        print(f"  {phase}: {len(normalized)} samples, "
              f"avg energy={fingerprints[phase]['avg_energy_kwh']:.2f} kWh")

    return fingerprints

def run_pelt_on_batch(power_signal):
    """Detect change points in power signal using PELT"""
    if len(power_signal) < 10:
        return []
    try:
        algo = rpt.Pelt(model='rbf', min_size=3, jump=1)
        algo.fit(power_signal.reshape(-1, 1))
        breakpoints = algo.predict(pen=3)
        return breakpoints[:-1]  # Remove last (= end of signal)
    except Exception:
        return []

def compute_energy_attribution(batch_ids):
    """Compute per-phase energy for all batches"""
    print("\nComputing phase energy attribution...")
    all_phase_power = {p: [] for p in PHASE_ORDER}
    energy_attribution = {}

    wb = openpyxl.load_workbook(PROCESS_FILE)

    for batch_id in batch_ids:
        sheet_name = f'Batch_{batch_id}'
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        phase_power = {p: [] for p in PHASE_ORDER}

        for r in range(2, ws.max_row + 1):
            phase = ws.cell(r, 3).value
            power = ws.cell(r, 10).value
            if phase and power is not None:
                power = max(0, float(power))
                if phase in phase_power:
                    phase_power[phase].append(power)

        # Energy per phase
        batch_energy = {}
        total_energy = 0
        for phase in PHASE_ORDER:
            vals = phase_power[phase]
            if vals:
                avg_power = np.mean(vals)
                duration = len(vals)
                energy = avg_power * duration / 60
                batch_energy[phase] = round(float(energy), 3)
                total_energy += energy
                all_phase_power[phase].append(np.array(vals))

        energy_attribution[batch_id] = {
            'phases': batch_energy,
            'total_kwh': round(total_energy, 2)
        }

    # Summary stats per phase
    phase_stats = {}
    for phase in PHASE_ORDER:
        energies = [
            energy_attribution[b]['phases'].get(phase, 0)
            for b in batch_ids
            if phase in energy_attribution.get(b, {}).get('phases', {})
        ]
        if energies:
            phase_stats[phase] = {
                'mean_kwh': round(float(np.mean(energies)), 2),
                'std_kwh': round(float(np.std(energies)), 2),
                'min_kwh': round(float(np.min(energies)), 2),
                'max_kwh': round(float(np.max(energies)), 2),
                'pct_of_total': round(float(np.mean(energies)) /
                    sum(s['mean_kwh'] for s in
                        {p: {'mean_kwh': float(np.mean([
                            energy_attribution[b]['phases'].get(p, 0)
                            for b in batch_ids
                            if p in energy_attribution.get(b, {}).get('phases', {})
                        ]))} for p in PHASE_ORDER}.values()
                        if s['mean_kwh'] > 0) * 100, 1)
            }

    print(f"\n  Phase energy breakdown (mean across {len(batch_ids)} batches):")
    for phase, stats in phase_stats.items():
        print(f"    {phase}: {stats['mean_kwh']:.2f} kWh "
              f"({stats['pct_of_total']:.1f}% of total)")

    return energy_attribution, phase_stats, all_phase_power

def update_supabase_energy(energy_attribution):
    """Update phase_sensors table with computed energy values"""
    print("\nUpdating Supabase phase energy values...")
    updated = 0
    for batch_id, data in energy_attribution.items():
        for phase, energy in data['phases'].items():
            supabase.table('phase_sensors').update(
                {'energy_kwh': energy}
            ).eq('batch_id', batch_id).eq('phase_name', phase).execute()
            updated += 1
    print(f"  Updated {updated} phase records")

if __name__ == '__main__':
    print("=" * 50)
    print("BatchMind — PELT Energy Attribution Script")
    print("=" * 50)

    batch_ids = get_all_batch_ids()
    print(f"Processing {len(batch_ids)} batches...")

    # Energy attribution
    energy_attribution, phase_stats, all_phase_power = compute_energy_attribution(batch_ids)

    # Build fingerprints
    fingerprints = compute_phase_fingerprints(all_phase_power)

    # PELT on a sample batch for demo
    print("\nRunning PELT change point detection on T014 (max energy)...")
    power_t014, phases_t014 = load_batch_power('T014')
    if power_t014 is not None:
        breakpoints = run_pelt_on_batch(power_t014)
        print(f"  Detected {len(breakpoints)} change points at minutes: {breakpoints}")

    print("\nRunning PELT on T038 (min energy)...")
    power_t038, phases_t038 = load_batch_power('T038')
    if power_t038 is not None:
        breakpoints_t038 = run_pelt_on_batch(power_t038)
        print(f"  Detected {len(breakpoints_t038)} change points at minutes: {breakpoints_t038}")

    # Update Supabase
    update_supabase_energy(energy_attribution)

    # Save outputs
    print("\nSaving outputs...")

    joblib.dump(energy_attribution, os.path.join(OUTPUT_DIR, 'energy_attribution.pkl'))
    print("  energy_attribution.pkl ✓")

    with open(os.path.join(OUTPUT_DIR, 'phase_stats.json'), 'w') as f:
        json.dump(phase_stats, f, indent=2)
    print("  phase_stats.json ✓")

    with open(os.path.join(OUTPUT_DIR, 'fingerprints.json'), 'w') as f:
        json.dump(fingerprints, f, indent=2)
    print("  fingerprints.json ✓")

    joblib.dump(
        {'T014': {'power': power_t014.tolist() if power_t014 is not None else [],
                  'phases': phases_t014},
         'T038': {'power': power_t038.tolist() if power_t038 is not None else [],
                  'phases': phases_t038}},
        os.path.join(OUTPUT_DIR, 'sample_power_curves.pkl')
    )
    print("  sample_power_curves.pkl ✓")

    print("\nPELT energy attribution complete.")